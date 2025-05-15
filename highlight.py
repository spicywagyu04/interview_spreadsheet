import json
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.hyperlink import Hyperlink
from copy import copy

# --- Helper Functions ---

def load_json_colleges_from_string(json_string):
    """Loads college names from a JSON string into a set."""
    try:
        json_data = json.loads(json_string) # Use json.loads() for string
        colleges = json_data.get('colleges', [])
        if not colleges:
            print("Warning: No colleges found in the provided JSON string or 'colleges' key is missing/empty.")
        return set(colleges)
    except json.JSONDecodeError:
        print("Error: Could not decode JSON from the provided string.")
        return None
    except Exception as e: # Catch other potential errors with string input
        print(f"An unexpected error occurred while processing the JSON string: {e}")
        return None

def create_cell_data_object(cell):
    """Creates a dictionary to store cell value, fill, font, and hyperlink."""
    return {
        "value": cell.value,
        "fill": copy(cell.fill) if cell.has_style and cell.fill else PatternFill(fill_type=None),
        "font": copy(cell.font) if cell.has_style and cell.font else Font(),
        "hyperlink": copy(cell.hyperlink) if cell.hyperlink else None,
    }

def extract_excel_data(sheet, start_row, college_name_col_idx, num_header_rows):
    """
    Extracts college data and headers from the Excel sheet,
    including cell values, fills, fonts, and hyperlinks.
    """
    existing_colleges_data = []
    existing_excel_college_names = set()
    header_cells_data = []

    num_columns = sheet.max_column
    if num_columns == 0 and sheet.max_row > 0:
        num_columns = len(sheet[1])
        
    for i in range(1, num_header_rows + 1):
        current_header_row = []
        for j in range(1, num_columns + 1):
            cell = sheet.cell(row=i, column=j)
            current_header_row.append(create_cell_data_object(cell))
        header_cells_data.append(current_header_row)

    for current_row_num in range(start_row, sheet.max_row + 1):
        college_name_cell_obj = sheet.cell(row=current_row_num, column=college_name_col_idx)
        college_name_cell_value = college_name_cell_obj.value

        if college_name_cell_value and isinstance(college_name_cell_value, str):
            college_name = college_name_cell_value.strip()
            if college_name:
                existing_excel_college_names.add(college_name)
                current_row_cell_objects = []
                for j in range(1, num_columns + 1):
                    cell = sheet.cell(row=current_row_num, column=j)
                    current_row_cell_objects.append(create_cell_data_object(cell))

                existing_colleges_data.append({
                    "name": college_name,
                    "cell_objects": current_row_cell_objects, # Store list of cell_data_objects
                    "is_new": False
                })
    return existing_colleges_data, existing_excel_college_names, header_cells_data, num_columns


def combine_and_prepare_data(existing_colleges, json_colleges_set, existing_excel_names, num_data_columns, college_name_col_idx):
    """Combines Excel and JSON data, marking colleges for highlighting and adding new ones."""
    all_processed_data = []

    for college in existing_colleges:
        is_in_json = college["name"] in json_colleges_set
        all_processed_data.append({
            "name": college["name"],
            "cell_objects": college["cell_objects"],
            "is_in_json": is_in_json,
            "is_new": False
        })

    for json_college_name in json_colleges_set:
        if json_college_name not in existing_excel_names:
            new_row_cell_objects = []
            for i in range(num_data_columns):
                col_idx_1_based = i + 1
                cell_val = None
                if col_idx_1_based == college_name_col_idx:
                    cell_val = json_college_name
                new_row_cell_objects.append({
                    "value": cell_val,
                    "fill": PatternFill(fill_type=None),
                    "font": Font(bold=True),
                    "hyperlink": None
                })

            all_processed_data.append({
                "name": json_college_name,
                "cell_objects": new_row_cell_objects,
                "is_in_json": True,
                "is_new": True
            })
    return all_processed_data

def write_data_to_new_sheet(workbook_new, sheet_name, header_cell_content, sorted_data, teal_fill, college_name_col_idx, num_header_rows):
    """Writes the processed and sorted data (including styles and hyperlinks) to a new sheet."""
    if sheet_name in workbook_new.sheetnames and len(workbook_new.sheetnames) > 1 and sheet_name != workbook_new.active.title:
        existing_sheet_to_remove = workbook_new[sheet_name]
        workbook_new.remove(existing_sheet_to_remove)
        new_sheet = workbook_new.create_sheet(title=sheet_name)
    elif sheet_name not in workbook_new.sheetnames :
        new_sheet = workbook_new.create_sheet(title=sheet_name)
    else:
        new_sheet = workbook_new[sheet_name]
        new_sheet.delete_rows(1, new_sheet.max_row + 1)

    for r_idx, header_row_cells in enumerate(header_cell_content):
        for c_idx, cell_data_obj in enumerate(header_row_cells):
            new_cell = new_sheet.cell(row=r_idx + 1, column=c_idx + 1)
            new_cell.value = cell_data_obj["value"]
            if cell_data_obj["fill"] and cell_data_obj["fill"].fill_type:
                new_cell.fill = cell_data_obj["fill"]
            if cell_data_obj["font"]:
                 new_cell.font = cell_data_obj["font"]
            if cell_data_obj["hyperlink"]:
                new_cell.hyperlink = cell_data_obj["hyperlink"]

    data_start_row = num_header_rows + 1
    for r_idx, college_entry in enumerate(sorted_data):
        current_excel_row = data_start_row + r_idx
        for c_idx, cell_data_obj in enumerate(college_entry['cell_objects']):
            actual_col_idx_1_based = c_idx + 1
            new_cell = new_sheet.cell(row=current_excel_row, column=actual_col_idx_1_based)
            new_cell.value = cell_data_obj["value"]

            if cell_data_obj["font"]:
                new_cell.font = cell_data_obj["font"]

            if actual_col_idx_1_based == college_name_col_idx and college_entry['is_in_json']:
                new_cell.fill = teal_fill
            elif cell_data_obj["fill"] and cell_data_obj["fill"].fill_type:
                new_cell.fill = cell_data_obj["fill"]
            else:
                new_cell.fill = PatternFill(fill_type=None)

            if cell_data_obj["hyperlink"]:
                new_cell.hyperlink = cell_data_obj["hyperlink"]

    return new_sheet

# --- Main Processing Function ---

def process_college_data_to_new_sheet(excel_filepath, json_college_names, output_filepath):
    """
    Loads college data, processes it, preserves original styles and hyperlinks,
    highlights names from JSON in teal, adds new names from JSON alphabetically,
    and writes the result to a new sheet in a new Excel file.
    """
    NEW_SHEET_NAME = "Sheet"
    TEAL_COLOR_HEX = "03fdfd"
    HEADER_ROWS = 3
    COLLEGE_NAME_COLUMN_IDX = 1
    
    json_colleges_set = load_json_colleges_from_string(json_college_names)
    if json_colleges_set is None:
        return False

    try:
        original_workbook = openpyxl.load_workbook(excel_filepath)
        original_sheet = original_workbook.active
    except FileNotFoundError:
        print(f"Error: Excel file not found at '{excel_filepath}'")
        return False
    except Exception as e:
        print(f"Error loading Excel file '{excel_filepath}': {e}")
        return False

    data_start_row = HEADER_ROWS + 1
    existing_colleges, existing_excel_names, header_cell_content, num_data_columns = extract_excel_data(
        original_sheet, data_start_row, COLLEGE_NAME_COLUMN_IDX, HEADER_ROWS
    )

    all_processed_data = combine_and_prepare_data(
        existing_colleges, json_colleges_set, existing_excel_names, num_data_columns, COLLEGE_NAME_COLUMN_IDX
    )

    sorted_colleges_data = sorted(all_processed_data, key=lambda x: (x['name'] is None, str(x['name']).lower()))

    output_workbook = openpyxl.Workbook()
    if NEW_SHEET_NAME in output_workbook.sheetnames and NEW_SHEET_NAME == output_workbook.active.title and len(output_workbook.sheetnames) ==1:
        pass
    elif "Sheet" in output_workbook.sheetnames and len(output_workbook.sheetnames) == 1 :
         output_workbook.remove(output_workbook["Sheet"])


    teal_fill = PatternFill(start_color=TEAL_COLOR_HEX, end_color=TEAL_COLOR_HEX, fill_type="solid")

    final_sheet = write_data_to_new_sheet(
        output_workbook, NEW_SHEET_NAME, header_cell_content, sorted_colleges_data,
        teal_fill, COLLEGE_NAME_COLUMN_IDX, HEADER_ROWS
    )
    if final_sheet:
        output_workbook.active = final_sheet

    try:
        output_workbook.save(output_filepath)
        print(f"\nSuccessfully processed data and saved to '{output_filepath}' in sheet '{NEW_SHEET_NAME}'.")
        print("Original styles (fills, hyperlinks) should be preserved, with teal override for matched college names.")
        return True
    except Exception as e:
        print(f"Error saving output workbook to '{output_filepath}': {e}")
        return False